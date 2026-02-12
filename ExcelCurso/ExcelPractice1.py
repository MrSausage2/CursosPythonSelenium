import openpyxl

book = openpyxl.load_workbook("E:\\Documentos\\Importantes\\Finanzas Actualizable.xlsx")
#Con esto se abre/ el documento de Excel y hay que asignarlo a un obj

sheet = book.active
Dict={}

sheet.cell(row=1, column=3).value = "Test text"
print(sheet.cell(row=1, column=3).value)
#Escribir y leer datos

print(sheet.max_row)#Retorna el numero de la ultima celda escrita en cualquier fila
print(sheet.max_column)#Retorna el numero de la ultima celda escrita en cualquier columna

for i in range(1, sheet.max_row + 1):
    if sheet.cell(row=i, column = 1).value =="Gybe":
        for j in range(1, sheet.max_column + 1):
                print(sheet.cell(row=i, column=j).value)



for x in range(1, sheet.max_row+1):
#Es importante que recuerdes que esto de arriba no define si la fila o
#columna se va a checar, solamente es para definir el rango del for
    if sheet.cell(row=x, column=8).value == "Pablo":
        print("Found pablo")
        for y in range(8, sheet.max_column + 1):
            Dict[sheet.cell(row=1 , column=y).value] = sheet.cell(row=x, column=y).value
            #Row 1 porque ahí está el título que en este caso es la key en el diccionario.

print(Dict)




#pip para trabajar con excel: https://pypi.org/project/openpyxl/