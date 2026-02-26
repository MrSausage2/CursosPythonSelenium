from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait
import openpyxl

def update_excel_data(file_path, searchTerm, colName, newValue):

    book = openpyxl.load_workbook(file_path)
    sheet = book.active
    ValueToModify={}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == colName:
            ValueToModify["col"]=i
        #Este for es para encontrar la columna donde se encuentra el titulo de precio

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == searchTerm:
                ValueToModify["row"] = i
        #Este for revisa todas las celdas hasta dar con el nombre de la fruta y guarda
        #el numero de la fila

    print(ValueToModify["row"])

    sheet.cell(row=ValueToModify["row"], column=ValueToModify["col"]).value=newValue
    book.save(file_path)


file_path = r"C:\Users\psanchez\Downloads\download.xlsx"#PC Amber
#file_path = r"E:\Documentos\Python practica\Webs de Prueba\download.xlsx"  # PC Casa
fruit_name = "Kivi"
newValue="800"
driver = webdriver.Chrome()
driver.implicitly_wait(6)
driver.get("https://rahulshettyacademy.com/upload-download-test/index.html")
driver.find_element(By.ID, "downloadButton").click()

update_excel_data(file_path, fruit_name, "price",newValue)

#Upload
file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

toast_locator=(By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
WebDriverWait(driver, 6).until(expected_conditions.visibility_of_element_located(toast_locator))

print(driver.find_element(*toast_locator).text)
priceColumn = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")


actual_price=(driver.find_element(By.XPATH, "//div[text()='"+fruit_name+"']/parent::div/parent::div/div[@id='cell-"+priceColumn+"-undefined']")).text
#//div[text()='"+fuit_name+"']/parent::div | /parent::div <- esto navega hacia el parent del elemento buscado, se pone la etiqueta en este caso "div"
#Recordar poner/no quitar -> ' ' <- cuando se ponga una variable, al final sigue siendo un elemento escrito en XPATH
#/parent::div/div[@id= | aquí se vuelve a navegar al padre "/parent::div"
#y después se navega al hijo "/div[@id="
#"priceColumn" toma el número de columna a través de un XPATH. De esta manera la búsqueda es dinámica.





print(fruit_name + " price: " + actual_price)
assert actual_price == newValue